#Encode

import arcpy, os, sys

GeodatabaseEntrada = r"X:\PRUEBAS\BackUp Corporativa Control\BK_27_07_2016.mdb"
ExcelFile=r"X:\PRUEBAS\BackUp Corporativa Control"
arcpy.env.workspace= GeodatabaseEntrada

ListaFeat = arcpy.ListFeatureClasses()
ListaTablas= arcpy.ListTables()
ListaDatasets= arcpy.ListDatasets("*","Feature")
ListaRelaciones=[]

desc = arcpy.Describe(GeodatabaseEntrada)

def listaCampos(Feat):
    lisCam=arcpy.ListFields(Feat)
    dicField={}
    for field in lisCam:
        if field.type=="Date":
            dicField[field.name]="Tipo Fecha"
        elif  field.type=="Double":
            dicField[field.name]="Tipo Numero Decimal"
        elif  field.type=="Integer":
            dicField[field.name]="Tipo Numero Entero Largo"
        elif  field.type=="SmallInteger":
            dicField[field.name]="Tipo Numero Entero Corto"
        elif  field.type=="String":
            dicField[field.name]="Texto de Longitud "+ str(field.length)
    return dicField
        #materias["lunes"] = [6103, 7540]



for child in desc.children:
    if child.datatype == "RelationshipClass":
        ListaRelaciones.append(child)

from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title="Relaciones"
i=1
for rel in ListaRelaciones:
    ws['A'+str(i)]=rel.name
    ws['B'+str(i)]=rel.dataType
    i=i+1
i=1
ws2 = wb.create_sheet("Tablas")
print "Creando Tablas"
for tab in ListaTablas:
    des= arcpy.Describe(tab)
    ws2['A'+str(i)]=des.name
    ws2['B'+str(i)]=des.dataType
    i=i+1

ws3 = wb.create_sheet("Featuare_Class_Ext")
i=1
print "Creando Featuare_Class"
for Feat in ListaFeat:
    des= arcpy.Describe(Feat)
    ws3['A'+str(i)]=des.name
    ws3['B'+str(i)]=des.dataType
    i=i+1

ws4 = wb.create_sheet("Datasets")
i=1
print "Creando Dataset"
for dat in ListaDatasets:
    des= arcpy.Describe(GeodatabaseEntrada+ os.sep+ dat)
    ws4['A'+str(i)]=des.name
    ws4['B'+str(i)]=des.dataType
    arcpy.env.workspace=GeodatabaseEntrada+ os.sep+ dat
    listaFeat2= arcpy.ListFeatureClasses()
    for fc in listaFeat2:
        i=i+1
        des= arcpy.Describe(fc)
        ws4['A'+str(i)]=des.name
        ws4['B'+str(i)]=des.dataType
    i=i+1

#Documentacion Campos
print "Creando Campos de Tablas"
for tab in ListaTablas:
    des= arcpy.Describe(GeodatabaseEntrada+os.sep+tab)
    wsfield= wb.create_sheet(des.name)
    i=1
    listaC= listaCampos(GeodatabaseEntrada+os.sep+tab)
    for nombre, tipo in listaC.items():
        wsfield['A1']="NOMBRE"
        wsfield['B1']="TIPO CAMPO"
        wsfield['C1']="DESCRIPCION"
        i=i+1
        wsfield['A'+str(i)]=nombre
        wsfield['B'+str(i)]=tipo

print "Creando Campos de Featuares"
for dat in ListaDatasets:
    des= arcpy.Describe(GeodatabaseEntrada+ os.sep+ dat)
    wsfield= wb.create_sheet(des.name)
    arcpy.env.workspace=GeodatabaseEntrada+ os.sep+ dat
    listaFeat2= arcpy.ListFeatureClasses()
    for fc in listaFeat2:
        listaC= listaCampos(GeodatabaseEntrada+os.sep+dat+os.sep+fc)
        des= arcpy.Describe(GeodatabaseEntrada+os.sep+dat+os.sep+fc)
        wsfield= wb.create_sheet(des.name)
        i=1
        for nombre, tipo in listaC.items():
            wsfield['A1']="NOMBRE"
            wsfield['B1']="TIPO CAMPO"
            wsfield['C1']="DESCRIPCION"
            i=i+1
            wsfield['A'+str(i)]=nombre
            wsfield['B'+str(i)]=tipo



wb.save(ExcelFile+os.sep+"sample.xlsx")